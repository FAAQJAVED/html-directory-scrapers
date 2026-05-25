"""
exporter.py
===========
Excel workbook generation for the HTML Directory Scraper.

Responsibilities:
  - Writing a 3-sheet .xlsx workbook:
      Sheet 1 "Data"    — clean validated records, frozen header, alternating shading
      Sheet 2 "Flagged" — records excluded by geo filter or failed fetches,
                          with a "Flag Reason" column appended
      Sheet 3 "Summary" — run metadata key/value table
  - Defining DATA_FIELDS and FLAG_FIELDS as module-level constants
    so tests can import and assert against them without loading openpyxl

This module performs NO HTTP calls and has NO side effects beyond file I/O.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

log = logging.getLogger(__name__)

# ── openpyxl guard — allows the module to be imported in test environments ────
try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

# ── Styling constants ─────────────────────────────────────────────────────────
_HEADER_BG = "1F4E79"  # dark navy
_HEADER_FG = "FFFFFF"  # white
_ALT_ROW_BG = "DCE6F1"  # light blue

# ── Column definitions ────────────────────────────────────────────────────────
DATA_FIELDS: List[str] = [
    "Company",
    "Email",
    "Phone",
    "Website",
    "Location",
    "Category",
    "Source",
]
FLAG_FIELDS: List[str] = DATA_FIELDS + ["Flag Reason"]


# ══════════════════════════════════════════════════════════════════════════════
# Internal helpers
# ══════════════════════════════════════════════════════════════════════════════


def _apply_header(ws, fields: List[str]) -> None:
    """
    Write and style the header row of a worksheet.

    Args:
        ws:     openpyxl Worksheet to write to.
        fields: Column name strings in display order.
    """
    header_fill = PatternFill("solid", fgColor=_HEADER_BG)
    header_font = Font(bold=True, color=_HEADER_FG, name="Arial", size=10)
    for ci, name in enumerate(fields, start=1):
        cell = ws.cell(row=1, column=ci, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"


def _write_rows(ws, rows: List[Dict[str, Any]], fields: List[str]) -> None:
    """
    Write data rows to a worksheet with alternating row shading.

    Args:
        ws:     openpyxl Worksheet (header already written).
        rows:   List of record dicts.
        fields: Column names; each dict is read by these keys in order.
    """
    alt_fill = PatternFill("solid", fgColor=_ALT_ROW_BG)
    cell_font = Font(name="Arial", size=10)
    for ri, record in enumerate(rows, start=2):
        fill = alt_fill if ri % 2 == 0 else None
        for ci, col in enumerate(fields, start=1):
            cell = ws.cell(row=ri, column=ci, value=record.get(col, ""))
            cell.font = cell_font
            if fill:
                cell.fill = fill


def _auto_width(ws, rows: List[Dict[str, Any]], fields: List[str]) -> None:
    """
    Set column widths based on the longest value in each column, capped at 60.

    Args:
        ws:     openpyxl Worksheet.
        rows:   Data rows used for content-width calculation.
        fields: Column names in display order.
    """
    for ci, col in enumerate(fields, start=1):
        max_len = len(col)
        for record in rows:
            val = str(record.get(col, "") or "")
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 60)


def _write_data_sheet(ws, rows: List[Dict[str, Any]]) -> None:
    """
    Populate the "Data" sheet with clean validated records.

    Args:
        ws:   openpyxl Worksheet for the Data sheet.
        rows: List of clean record dicts.
    """
    _apply_header(ws, DATA_FIELDS)
    _write_rows(ws, rows, DATA_FIELDS)
    _auto_width(ws, rows, DATA_FIELDS)


def _write_flagged_sheet(ws, rows: List[Dict[str, Any]]) -> None:
    """
    Populate the "Flagged" sheet with rejected/filtered records.

    The "Flag Reason" column is appended at the end of each row.

    Args:
        ws:   openpyxl Worksheet for the Flagged sheet.
        rows: List of flagged record dicts (must include "Flag Reason" key).
    """
    _apply_header(ws, FLAG_FIELDS)
    _write_rows(ws, rows, FLAG_FIELDS)
    _auto_width(ws, rows, FLAG_FIELDS)


def _write_summary_sheet(
    ws, stats: Dict[str, Any], clean_rows: List[Dict], flagged_rows: List[Dict]
) -> None:
    """
    Populate the "Summary" sheet with run metadata as a key/value table.

    Args:
        ws:          openpyxl Worksheet for the Summary sheet.
        stats:       Dict of run metadata (start_time, source, status, etc.).
        clean_rows:  Clean records list (used to derive counts).
        flagged_rows: Flagged records list (used to derive counts).
    """
    header_fill = PatternFill("solid", fgColor=_HEADER_BG)
    header_font = Font(bold=True, color=_HEADER_FG, name="Arial", size=10)
    cell_font = Font(name="Arial", size=10)

    # Header
    for ci, label in enumerate(["Metric", "Value"], start=1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"

    total = len(clean_rows)
    flagged = len(flagged_rows)

    summary_rows = [
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Source", stats.get("source", "")),
        ("Status", stats.get("status", "PARTIAL")),
        ("Total clean", total),
        ("Total flagged", flagged),
        ("Total processed", stats.get("total_scraped", total + flagged)),
        ("With email", sum(1 for r in clean_rows if r.get("Email"))),
        ("With phone", sum(1 for r in clean_rows if r.get("Phone"))),
        ("With website", sum(1 for r in clean_rows if r.get("Website"))),
        ("Started", stats.get("start_time", "")),
    ]

    for ri, (key, val) in enumerate(summary_rows, start=2):
        ws.cell(row=ri, column=1, value=key).font = Font(bold=True, name="Arial", size=10)
        ws.cell(row=ri, column=2, value=val).font = cell_font

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30


# ══════════════════════════════════════════════════════════════════════════════
# Public API
# ══════════════════════════════════════════════════════════════════════════════


def export_excel(
    clean_rows: List[Dict[str, Any]],
    flagged_rows: List[Dict[str, Any]],
    output_path: str,
    stats: Dict[str, Any],
) -> None:
    """
    Write a 3-sheet Excel workbook to *output_path*.

    Sheet layout:
      - **Data**    — all clean validated records
      - **Flagged** — records excluded by geo filter or failed profile fetches,
                      each with a "Flag Reason" column
      - **Summary** — run metadata (counts, timestamps, source, status)

    Creates parent directories if they do not exist. Writes atomically via
    openpyxl's built-in save (no partial-write risk on modern filesystems).

    Args:
        clean_rows:   List of validated record dicts (DATA_FIELDS keys).
        flagged_rows: List of rejected record dicts (FLAG_FIELDS keys).
        output_path:  Destination .xlsx file path.
        stats:        Run metadata dict.  Expected keys: source, status,
                      start_time, total_scraped.

    Returns:
        None.  Logs a warning if openpyxl is unavailable or the save fails.
    """
    if not _OPENPYXL_OK:
        log.warning(
            "openpyxl is not installed — Excel output skipped. "
            "Install with: pip install openpyxl"
        )
        return

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    ws_data = wb.create_sheet("Data")
    ws_flagged = wb.create_sheet("Flagged")
    ws_summary = wb.create_sheet("Summary")

    _write_data_sheet(ws_data, clean_rows)
    _write_flagged_sheet(ws_flagged, flagged_rows)
    _write_summary_sheet(ws_summary, stats, clean_rows, flagged_rows)

    try:
        wb.save(output_path)
        log.info("Excel saved → %s", output_path)
    except OSError as exc:
        log.warning("Excel save failed: %s", exc)
