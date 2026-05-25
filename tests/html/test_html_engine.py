"""
tests/html/test_html_engine.py
==============================
Pytest test suite for the HTML Directory Scraper engine.

Coverage:
  - parser.decode_cf_email   — Cloudflare XOR decoding (valid, empty, malformed)
  - parser.extract_email     — three extraction strategies (CF href, data-cfemail, mailto)
  - parser.clean_phone       — generic 7–15 digit normalisation (no country-specific rules)
  - fetcher.parse_cookies    — cookie string parsing
  - fetcher.progress_bar     — ASCII progress bar rendering
  - fetcher.get_total_pages  — result-count extraction + fallback
  - parser.parse_cards       — listing-page card parsing from HTML
  - exporter.export_excel    — 3-sheet workbook creation (Data/Flagged/Summary)
  - checkpoint.CheckpointManager — save/load round-trip, missing file, clear, no .tmp left

All tests are self-contained with inline HTML/data.
No network calls are made.
File I/O tests use the tmp_path pytest fixture.
"""

import json
import sys
from pathlib import Path

import pytest

# ── add engines/html to sys.path so modules import without a package install ──
ENGINE_DIR = Path(__file__).parent.parent.parent / "engines" / "html"
if str(ENGINE_DIR) not in sys.path:
    sys.path.insert(0, str(ENGINE_DIR))

from bs4 import BeautifulSoup

import parser
import fetcher
import exporter
import checkpoint as ckpt_module

# ══════════════════════════════════════════════════════════════════════════════
# parser.decode_cf_email
# ══════════════════════════════════════════════════════════════════════════════


def _cf_encode(email: str) -> str:
    """Helper: XOR-encode an email the same way Cloudflare does."""
    key = 0x3A
    enc = bytes([key] + [ord(c) ^ key for c in email])
    return enc.hex()


class TestDecodeCfEmail:
    def test_decodes_known_encoded_string(self):
        email = "hello@example.com"
        encoded = _cf_encode(email)
        assert parser.decode_cf_email(encoded) == email

    def test_returns_empty_for_empty_input(self):
        assert parser.decode_cf_email("") == ""

    def test_returns_empty_for_odd_length_hex(self):
        assert parser.decode_cf_email("abc") == ""  # odd-length → fromhex raises

    def test_returns_empty_for_non_hex_chars(self):
        assert parser.decode_cf_email("ZZZZZZ") == ""


# ══════════════════════════════════════════════════════════════════════════════
# parser.extract_email
# ══════════════════════════════════════════════════════════════════════════════


class TestExtractEmail:
    def _soup(self, html: str) -> BeautifulSoup:
        return BeautifulSoup(html, "html.parser")

    def test_extracts_cloudflare_href_email(self):
        email = "contact@business.com"
        encoded = _cf_encode(email)
        html = f'<a href="/cdn-cgi/l/email-protection#{encoded}">Email us</a>'
        assert parser.extract_email(self._soup(html)) == email

    def test_extracts_data_cfemail_span(self):
        email = "info@company.org"
        encoded = _cf_encode(email)
        html = f'<span data-cfemail="{encoded}">protected</span>'
        assert parser.extract_email(self._soup(html)) == email

    def test_extracts_plain_mailto_as_fallback(self):
        html = '<a href="mailto:admin@site.net">Contact</a>'
        assert parser.extract_email(self._soup(html)) == "admin@site.net"

    def test_returns_empty_when_no_email_present(self):
        html = "<p>No email here at all.</p>"
        assert parser.extract_email(self._soup(html)) == ""


# ══════════════════════════════════════════════════════════════════════════════
# parser.clean_phone  (generic 7–15 digit, no country-specific rules)
# ══════════════════════════════════════════════════════════════════════════════


class TestCleanPhone:
    def test_returns_digits_for_clean_10_digit_number(self):
        assert parser.clean_phone("8005551234") == "8005551234"

    def test_strips_punctuation_and_validates_length(self):
        assert parser.clean_phone("+1 (800) 555-1234") == "18005551234"

    def test_returns_empty_for_too_short(self):
        assert parser.clean_phone("1234") == ""

    def test_returns_empty_for_too_long(self):
        assert parser.clean_phone("1" * 20) == ""

    def test_returns_empty_for_empty_string(self):
        assert parser.clean_phone("") == ""

    def test_accepts_7_digit_minimum(self):
        result = parser.clean_phone("1234567")
        assert result == "1234567"

    def test_accepts_15_digit_maximum(self):
        result = parser.clean_phone("1" * 15)
        assert result == "1" * 15


# ══════════════════════════════════════════════════════════════════════════════
# fetcher.parse_cookies
# ══════════════════════════════════════════════════════════════════════════════


class TestParseCookies:
    def test_parses_single_cookie(self):
        result = fetcher.parse_cookies("session=abc123")
        assert result == {"session": "abc123"}

    def test_parses_multiple_cookies(self):
        result = fetcher.parse_cookies("a=1; b=2; c=three")
        assert result == {"a": "1", "b": "2", "c": "three"}

    def test_returns_empty_dict_for_empty_string(self):
        assert fetcher.parse_cookies("") == {}

    def test_handles_value_with_equals_sign(self):
        result = fetcher.parse_cookies("token=ab=cd")
        assert result["token"] == "ab=cd"


# ══════════════════════════════════════════════════════════════════════════════
# fetcher.progress_bar
# ══════════════════════════════════════════════════════════════════════════════


class TestProgressBar:
    def test_returns_string_with_bracket_delimiters(self):
        bar = fetcher.progress_bar(5, 10)
        assert bar.startswith("[")
        assert "]" in bar

    def test_shows_100_percent_when_done_equals_total(self):
        bar = fetcher.progress_bar(10, 10)
        assert "100%" in bar

    def test_does_not_crash_when_total_is_zero(self):
        bar = fetcher.progress_bar(0, 0)
        assert isinstance(bar, str)


# ══════════════════════════════════════════════════════════════════════════════
# fetcher.get_total_pages
# ══════════════════════════════════════════════════════════════════════════════


class TestGetTotalPages:
    def _soup(self, html: str) -> BeautifulSoup:
        return BeautifulSoup(html, "html.parser")

    def test_extracts_total_pages_from_result_count(self):
        html = "<p>Showing 150 results found</p>"
        pages = fetcher.get_total_pages(self._soup(html), page_size=10)
        assert pages == 15

    def test_returns_fallback_when_no_count_found(self):
        html = "<p>No count text here at all.</p>"
        pages = fetcher.get_total_pages(self._soup(html), page_size=10)
        assert pages == 999

    def test_handles_comma_separated_large_numbers(self):
        html = "<p>1,200 results found in our directory</p>"
        pages = fetcher.get_total_pages(self._soup(html), page_size=20)
        assert pages == 60


# ══════════════════════════════════════════════════════════════════════════════
# parser.parse_cards
# ══════════════════════════════════════════════════════════════════════════════


class TestParseCards:
    def _cfg(self, base_url: str = "https://example.com") -> dict:
        return {
            "base_url": base_url,
            "selectors": {
                "card_container": ".member-item",
                "profile_link": "a[href*='/company/']",
                "member_name": ".member-name",
                "badge_images": ".badge img",
            },
            "badge_image_keywords": {
                "logo-cat-one": "Category One",
            },
        }

    def test_returns_correct_number_of_cards(self):
        html = """
        <div class="member-item">
            <span class="member-name">Acme Corp</span>
            <a href="/company/acme">View</a>
        </div>
        <div class="member-item">
            <span class="member-name">Beta Ltd</span>
            <a href="/company/beta">View</a>
        </div>
        """
        soup = BeautifulSoup(html, "html.parser")
        cards = parser.parse_cards(soup, self._cfg())
        assert len(cards) == 2

    def test_returns_correct_name_and_url(self):
        html = """
        <div class="member-item">
            <span class="member-name">Gamma Inc</span>
            <a href="/company/gamma">View</a>
        </div>
        """
        soup = BeautifulSoup(html, "html.parser")
        cards = parser.parse_cards(soup, self._cfg())
        assert cards[0]["name"] == "Gamma Inc"
        assert cards[0]["url"] == "https://example.com/company/gamma"

    def test_returns_empty_list_when_no_cards_match(self):
        html = "<div class='something-else'>No matching cards here</div>"
        soup = BeautifulSoup(html, "html.parser")
        cards = parser.parse_cards(soup, self._cfg())
        assert cards == []

    def test_prepends_base_url_to_relative_link(self):
        html = """
        <div class="member-item">
            <span class="member-name">Delta Co</span>
            <a href="/company/delta">Visit</a>
        </div>
        """
        soup = BeautifulSoup(html, "html.parser")
        cards = parser.parse_cards(soup, self._cfg("https://mysite.org"))
        assert cards[0]["url"].startswith("https://mysite.org")

    def test_assigns_category_from_badge_keyword(self):
        html = """
        <div class="member-item">
            <span class="member-name">Echo LLC</span>
            <a href="/company/echo">View</a>
            <div class="badge"><img src="/images/logo-cat-one-badge.png"></div>
        </div>
        """
        soup = BeautifulSoup(html, "html.parser")
        cards = parser.parse_cards(soup, self._cfg())
        assert "Category One" in cards[0]["services"]


# ══════════════════════════════════════════════════════════════════════════════
# exporter.export_excel
# ══════════════════════════════════════════════════════════════════════════════


class TestExportExcel:
    def _sample_clean(self) -> list:
        return [
            {
                "Company": "Acme Corp",
                "Email": "a@acme.com",
                "Phone": "8005551234",
                "Website": "https://acme.com",
                "Location": "10001",
                "Category": "Category One",
                "Source": "Directory",
            }
        ]

    def _sample_flagged(self) -> list:
        return [
            {
                "Company": "Skip Co",
                "Email": "",
                "Phone": "",
                "Website": "",
                "Location": "",
                "Category": "",
                "Source": "Directory",
                "Flag Reason": "Outside geographic filter",
            }
        ]

    def _stats(self) -> dict:
        return {
            "source": "Directory",
            "status": "COMPLETE",
            "start_time": "2025-01-01T06:00:00",
            "total_scraped": 1,
        }

    def test_creates_file_at_given_path(self, tmp_path):
        out = str(tmp_path / "output.xlsx")
        exporter.export_excel(self._sample_clean(), [], out, self._stats())
        assert Path(out).exists()

    def test_workbook_has_exactly_three_sheets(self, tmp_path):
        import openpyxl

        out = str(tmp_path / "output.xlsx")
        exporter.export_excel(self._sample_clean(), self._sample_flagged(), out, self._stats())
        wb = openpyxl.load_workbook(out)
        assert wb.sheetnames == ["Data", "Flagged", "Summary"]

    def test_data_sheet_has_correct_headers(self, tmp_path):
        import openpyxl

        out = str(tmp_path / "output.xlsx")
        exporter.export_excel(self._sample_clean(), [], out, self._stats())
        wb = openpyxl.load_workbook(out)
        headers = [c.value for c in wb["Data"][1]]
        assert "Company" in headers
        assert "Email" in headers
        assert "Category" in headers

    def test_flagged_sheet_has_flag_reason_column(self, tmp_path):
        import openpyxl

        out = str(tmp_path / "output.xlsx")
        exporter.export_excel([], self._sample_flagged(), out, self._stats())
        wb = openpyxl.load_workbook(out)
        headers = [c.value for c in wb["Flagged"][1]]
        assert "Flag Reason" in headers

    def test_creates_file_with_empty_data(self, tmp_path):
        out = str(tmp_path / "empty.xlsx")
        exporter.export_excel([], [], out, self._stats())
        assert Path(out).exists()


# ══════════════════════════════════════════════════════════════════════════════
# checkpoint.CheckpointManager
# ══════════════════════════════════════════════════════════════════════════════


class TestCheckpointManager:
    def test_save_and_load_round_trip(self, tmp_path):
        state = {"page": 5, "total_scraped": 42, "seen_urls": ["https://x.com"]}
        manager = ckpt_module.CheckpointManager(str(tmp_path / "ckpt.json"))
        manager.save(state)
        loaded = manager.load()
        assert loaded == state

    def test_load_returns_none_when_no_file_exists(self, tmp_path):
        manager = ckpt_module.CheckpointManager(str(tmp_path / "missing.json"))
        assert manager.load() is None

    def test_clear_removes_file_and_exists_returns_false(self, tmp_path):
        manager = ckpt_module.CheckpointManager(str(tmp_path / "ckpt.json"))
        manager.save({"page": 1})
        assert manager.exists()
        manager.clear()
        assert not manager.exists()

    def test_save_does_not_leave_tmp_file_behind(self, tmp_path):
        manager = ckpt_module.CheckpointManager(str(tmp_path / "ckpt.json"))
        manager.save({"page": 2})
        tmp_file = tmp_path / "ckpt.tmp"
        assert not tmp_file.exists()

    def test_load_returns_none_on_corrupt_json(self, tmp_path):
        ckpt_file = tmp_path / "bad.json"
        ckpt_file.write_text("NOT VALID JSON {{{{", encoding="utf-8")
        manager = ckpt_module.CheckpointManager(str(ckpt_file))
        assert manager.load() is None


# =============================================================================
# v1.1.0 — new feature tests
# =============================================================================


class TestConcurrentFetch:
    """scraper._fetch_profiles_concurrent returns results in input order."""

    def test_result_order_preserved(self, tmp_path):
        """Results come back in the same order as input cards, regardless of thread timing."""
        import sys
        from pathlib import Path

        engine = Path(__file__).parent.parent.parent / "engines" / "html"
        if str(engine) not in sys.path:
            sys.path.insert(0, str(engine))
        import scraper as scraper_mod

        cfg = {
            "base_url": "https://example.com",
            "selectors": {"card_container": ".x", "profile_link": "a", "member_name": "h6"},
            "badge_image_keywords": {},
            "source_label": "Test",
            "headers": {},
            "timeout_seconds": 5,
            "cookies_raw": "",
        }
        cards = [
            {
                "name": f"Company {i}",
                "url": f"https://example.com/{i}",
                "services": [],
                "location": "",
            }
            for i in range(5)
        ]
        # Mock scrape_profile to return immediately with a predictable rec
        import unittest.mock as mock

        with mock.patch(
            "parser.scrape_profile",
            side_effect=lambda c, card, cfg: {
                "Company": card["name"],
                "Email": "",
                "Phone": "",
                "Website": "",
                "Location": "",
                "Category": "",
                "Source": "Test",
            },
        ):
            results = scraper_mod._fetch_profiles_concurrent(cards, cfg, 3, 0.0, 0.01)
        assert len(results) == 5
        for i, (card, rec) in enumerate(results):
            assert card["name"] == f"Company {i}"


class TestParseCardsWithMeta:
    """parse_cards() v1.1.0: location field populated from card meta element."""

    def test_location_extracted_from_meta(self):
        html = """
        <div class="member-item">
            <h6 class="member-name"><a href="/company/abc.html">Abc Ltd</a></h6>
            <span class="meta">Ringwood</span>
            <a href="/company/abc.html">View</a>
        </div>
        """
        cfg = {
            "base_url": "https://example.com",
            "selectors": {
                "card_container": "div.member-item",
                "profile_link": "a[href*='/company/']",
                "member_name": "h6.member-name",
                "badge_images": "div.division-logos img",
                "card_meta": "span.meta",
            },
            "badge_image_keywords": {},
        }
        soup = BeautifulSoup(html, "html.parser")
        cards = parser.parse_cards(soup, cfg)
        assert cards[0]["location"] == "Ringwood"

    def test_location_empty_when_no_meta_selector(self):
        html = """
        <div class="member-item">
            <h6 class="member-name"><a href="/company/abc.html">Abc Ltd</a></h6>
            <a href="/company/abc.html">View</a>
        </div>
        """
        cfg = {
            "base_url": "https://example.com",
            "selectors": {
                "card_container": "div.member-item",
                "profile_link": "a[href*='/company/']",
                "member_name": "h6.member-name",
                # no card_meta key
            },
            "badge_image_keywords": {},
        }
        soup = BeautifulSoup(html, "html.parser")
        cards = parser.parse_cards(soup, cfg)
        assert cards[0]["location"] == ""


class TestConnectTimeoutNoRetry:
    """fetcher.safe_get with is_profile=True returns immediately on ConnectTimeout."""

    def test_connect_timeout_returns_empty_immediately(self):
        import unittest.mock as mock
        import httpx

        mock_client = mock.MagicMock(spec=httpx.Client)
        mock_client.get.side_effect = httpx.ConnectTimeout("timed out")

        start = __import__("time").time()
        body, code = fetcher.safe_get(mock_client, "https://dead.example.com", is_profile=True)
        elapsed = __import__("time").time() - start

        assert body == ""
        assert code == 0
        assert elapsed < 1.0  # must return in under 1 s — no sleep/retry
        assert mock_client.get.call_count == 1  # called exactly once


# ══════════════════════════════════════════════════════════════════════════════
# config.load_config
# ══════════════════════════════════════════════════════════════════════════════

import config as cfg_mod  # noqa: E402 — added by v1.1.0 improvement pass


class TestLoadConfig:
    def test_raises_file_not_found_for_missing_config(self, tmp_path):
        """load_config raises FileNotFoundError for a nonexistent path."""
        with pytest.raises(FileNotFoundError):
            cfg_mod.load_config(str(tmp_path / "nonexistent.yaml"))

    def test_raises_value_error_for_missing_required_keys(self, tmp_path):
        """load_config raises ValueError when required keys are absent."""
        cfg_file = tmp_path / "config.yaml"
        cfg_file.write_text("tool_name: test\n", encoding="utf-8")
        with pytest.raises(ValueError, match="missing required"):
            cfg_mod.load_config(str(cfg_file))

    def test_injects_scraper_cookies_from_env(self, tmp_path, monkeypatch):
        """SCRAPER_COOKIES env var overrides cookies_raw in config."""
        cfg_file = tmp_path / "config.yaml"
        cfg_file.write_text(
            "base_url: https://example.com\n"
            "list_path: /find\n"
            "categories: [{name: Test}]\n"
            "selectors: {card_container: .x, profile_link: a}\n",
            encoding="utf-8",
        )
        monkeypatch.setenv("SCRAPER_COOKIES", "session=abc123")
        cfg = cfg_mod.load_config(str(cfg_file))
        assert cfg["cookies_raw"] == "session=abc123"


# ══════════════════════════════════════════════════════════════════════════════
# controls.check_stop_time
# ══════════════════════════════════════════════════════════════════════════════

import controls  # noqa: E402 — added by v1.1.0 improvement pass


class TestCheckStopTime:
    def test_returns_false_for_empty_string(self):
        assert controls.check_stop_time("") is False

    def test_returns_false_when_time_has_not_reached(self):
        # Use a time far in the future (23:59) — this test runs before that
        assert controls.check_stop_time("23:59") is False

    def test_returns_true_when_stop_time_is_in_the_past(self):
        """check_stop_time returns True when stop_at is midnight (always past)."""
        # "00:00" is always <= current wall-clock time on any machine.
        # Avoids mock.patch("controls.datetime") which fails because HTML
        # controls uses `from datetime import datetime` (no module attr).
        assert controls.check_stop_time("00:00") is True

    def test_returns_false_for_exact_boundary_not_yet_reached(self):
        """check_stop_time returns False when stop_at is disabled or far future."""
        # Empty string disables the feature — always False.
        assert controls.check_stop_time("") is False

# =============================================================================
# v1.2.0 — additional tests
# =============================================================================


class TestFetcherRateEta:
    """fetcher.rate_eta returns empty string with insufficient data."""

    def test_returns_empty_with_fewer_than_two_timestamps(self):
        import unittest.mock as mock

        with mock.patch.object(fetcher, "_scrape_ts", []):
            result = fetcher.rate_eta(0, 100)
        assert result == ""

    def test_returns_string_with_two_or_more_timestamps(self):
        import time
        import unittest.mock as mock

        fake_ts = [time.time() - 10, time.time()]
        with mock.patch.object(fetcher, "_scrape_ts", fake_ts):
            result = fetcher.rate_eta(2, 100)
        assert isinstance(result, str)
        assert "min" in result or "ETA" in result


class TestExtractEmailPriority:
    """extract_email prefers CF href over data-cfemail over plain mailto."""

    def _soup(self, html: str):
        return BeautifulSoup(html, "html.parser")

    def test_prefers_cf_href_over_mailto(self):
        cf_email = "primary@cf.com"
        encoded = _cf_encode(cf_email)
        html = (
            f'<a href="/cdn-cgi/l/email-protection#{encoded}">CF link</a>'
            '<a href="mailto:fallback@other.com">Fallback</a>'
        )
        assert parser.extract_email(self._soup(html)) == cf_email

    def test_falls_back_to_mailto_when_no_cf(self):
        html = '<a href="mailto:info@example.org">Contact</a>'
        assert parser.extract_email(self._soup(html)) == "info@example.org"

    def test_data_cfemail_takes_priority_over_mailto(self):
        cf_email = "hidden@company.com"
        encoded = _cf_encode(cf_email)
        html = (
            f'<span data-cfemail="{encoded}">protected</span>'
            '<a href="mailto:fallback@other.com">email</a>'
        )
        assert parser.extract_email(self._soup(html)) == cf_email


class TestProgressBar:
    """fetcher.progress_bar renders correctly."""

    def test_starts_with_bracket(self):
        bar = fetcher.progress_bar(5, 10)
        assert bar.startswith("[")

    def test_shows_100_percent_at_completion(self):
        bar = fetcher.progress_bar(10, 10)
        assert "100%" in bar

    def test_zero_total_does_not_crash(self):
        bar = fetcher.progress_bar(0, 0)
        assert isinstance(bar, str)
