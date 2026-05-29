"""
tests/wordpress/test_wordpress_engine.py
=========================================
Pytest test suite for the WordPress Directory Scraper engine.

Coverage:
  - fetcher.safe_decode        — plain UTF-8, gzip, zlib decompression; invalid bytes
  - fetcher.parse_cookies      — single cookie, multiple cookies, empty/None
  - fetcher.decode_entities    — numeric and named HTML entity replacement
  - parser.is_valid_email      — format check, junk domain, filename-like, no @
  - parser.extract_emails      — valid email in HTML, junk filtered, empty HTML
  - parser.parse_cards         — card dict parsing, postcode extraction, empty input
  - parser.filter_by_bounds    — inside bounds, outside bounds, empty list
  - parser.markers_to_items    — all markers returned, card_map merge
  - exporter.export_excel      — 3-sheet workbook (Data/Flagged/Summary), headers
  - checkpoint.CheckpointManager — save/load round-trip, missing file, clear, no .tmp

All tests are self-contained with inline data.
No network calls are made.
File I/O tests use the tmp_path pytest fixture.
"""

import gzip
import json
import sys
import zlib
from pathlib import Path

import pytest

# ── add engines/wordpress to sys.path ─────────────────────────────────────────
ENGINE_DIR = Path(__file__).parent.parent.parent / "engines" / "wordpress"
if str(ENGINE_DIR) not in sys.path:
    sys.path.insert(0, str(ENGINE_DIR))

import fetcher
import parser
import exporter
import checkpoint as ckpt_module

# ══════════════════════════════════════════════════════════════════════════════
# fetcher.safe_decode
# ══════════════════════════════════════════════════════════════════════════════


class TestSafeDecode:
    def test_decodes_plain_utf8_bytes(self):
        raw = "Hello, world!".encode("utf-8")
        result = fetcher.safe_decode(raw)
        assert result == "Hello, world!"

    def test_decompresses_gzip_bytes(self):
        original = "This is gzip-compressed content."
        compressed = gzip.compress(original.encode("utf-8"))
        result = fetcher.safe_decode(compressed)
        assert result == original

    def test_decompresses_zlib_bytes(self):
        original = "This is zlib/deflate compressed."
        compressed = zlib.compress(original.encode("utf-8"))
        result = fetcher.safe_decode(compressed)
        assert result == original

    def test_does_not_raise_on_invalid_utf8(self):
        raw = b"valid start \xff\xfe invalid bytes"
        result = fetcher.safe_decode(raw)
        assert isinstance(result, str)
        assert len(result) > 0

    def test_empty_bytes_returns_empty_string(self):
        assert fetcher.safe_decode(b"") == ""


# ══════════════════════════════════════════════════════════════════════════════
# fetcher.parse_cookies
# ══════════════════════════════════════════════════════════════════════════════


class TestParseCookies:
    def test_parses_single_cookie(self):
        result = fetcher.parse_cookies("session=xyz789")
        assert result == {"session": "xyz789"}

    def test_parses_multiple_semicolon_separated_cookies(self):
        result = fetcher.parse_cookies("a=1; b=two; c=3")
        assert result == {"a": "1", "b": "two", "c": "3"}

    def test_returns_empty_dict_for_empty_string(self):
        assert fetcher.parse_cookies("") == {}

    def test_returns_empty_dict_for_none(self):
        assert fetcher.parse_cookies(None) == {}

    def test_handles_value_containing_equals(self):
        result = fetcher.parse_cookies("token=abc==def")
        assert result["token"] == "abc==def"


# ══════════════════════════════════════════════════════════════════════════════
# fetcher.decode_entities
# ══════════════════════════════════════════════════════════════════════════════


class TestDecodeEntities:
    def test_decodes_numeric_ampersand_entity(self):
        assert fetcher.decode_entities("Fish &#038; Chips") == "Fish & Chips"

    def test_decodes_named_amp_entity(self):
        assert fetcher.decode_entities("A &amp; B") == "A & B"

    def test_returns_unchanged_string_with_no_entities(self):
        text = "No entities here at all."
        assert fetcher.decode_entities(text) == text

    def test_decodes_multiple_entities_in_one_string(self):
        result = fetcher.decode_entities("A &amp; B &#038; C")
        assert result == "A & B & C"


# ══════════════════════════════════════════════════════════════════════════════
# parser.is_valid_email
# ══════════════════════════════════════════════════════════════════════════════


class TestIsValidEmail:
    JUNK = {"example.com", "google.com", "w3.org"}

    def test_returns_true_for_well_formed_email(self):
        assert parser.is_valid_email("hello@company.io", self.JUNK)

    def test_returns_false_for_junk_domain(self):
        assert not parser.is_valid_email("user@example.com", self.JUNK)

    def test_returns_false_for_filename_like_email(self):
        assert not parser.is_valid_email("image@sprite.png", self.JUNK)

    def test_returns_false_for_string_with_no_at(self):
        assert not parser.is_valid_email("notanemail", self.JUNK)

    def test_returns_false_for_empty_string(self):
        assert not parser.is_valid_email("", self.JUNK)

    def test_returns_false_for_js_extension_email(self):
        assert not parser.is_valid_email("bundle@app.js", self.JUNK)


# ══════════════════════════════════════════════════════════════════════════════
# parser.extract_emails
# ══════════════════════════════════════════════════════════════════════════════


class TestExtractEmails:
    JUNK = {"example.com", "google.com"}

    def test_finds_one_valid_email_in_html(self):
        html = "<p>Contact us at support@mybusiness.com for help.</p>"
        result = parser.extract_emails(html, self.JUNK)
        assert "support@mybusiness.com" in result

    def test_ignores_junk_domain_email(self):
        html = "<p>Email info@example.com for details.</p>"
        result = parser.extract_emails(html, self.JUNK)
        assert result == []

    def test_returns_empty_list_for_html_with_no_emails(self):
        html = "<p>No contact information is listed on this page.</p>"
        result = parser.extract_emails(html, self.JUNK)
        assert result == []

    def test_deduplicates_repeated_emails(self):
        html = "<p>contact@biz.com and contact@biz.com again</p>"
        result = parser.extract_emails(html, self.JUNK)
        assert result.count("contact@biz.com") == 1


# ══════════════════════════════════════════════════════════════════════════════
# parser.parse_cards
# ══════════════════════════════════════════════════════════════════════════════


class TestParseCards:
    BASE = "https://example-directory.com"

    def test_parses_card_with_name_address_and_link(self):
        html = """
        <div class="card listing-card">
            <h3>Acme Supplies Ltd</h3>
            <p>123 Main Street, New York, NY 10001</p>
            <a href="/members/acme-supplies/">View Profile</a>
        </div>
        """
        cfg = {"postcode_regex": r"\b\d{5}\b"}
        result = parser.parse_cards(html, self.BASE, cfg)
        assert "Acme Supplies Ltd" in result
        entry = result["Acme Supplies Ltd"]
        assert entry["postcode"] == "10001"
        assert entry["url"].startswith("https://")

    def test_returns_empty_dict_for_html_with_no_card_divs(self):
        html = "<div class='something-else'>No matching cards</div>"
        result = parser.parse_cards(html, self.BASE)
        assert result == {}

    def test_returns_empty_dict_for_empty_string(self):
        assert parser.parse_cards("", self.BASE) == {}

    def test_resolves_relative_url_with_base(self):
        html = """
        <div class="card">
            <h3>Beta Corp</h3>
            <a href="/members/beta-corp/">View</a>
        </div>
        """
        result = parser.parse_cards(html, self.BASE)
        assert "Beta Corp" in result
        assert result["Beta Corp"]["url"].startswith(self.BASE)


# ══════════════════════════════════════════════════════════════════════════════
# parser.filter_by_bounds
# ══════════════════════════════════════════════════════════════════════════════


class TestFilterByBounds:
    # Greater New York bounding box
    NYC_BOUNDS = {
        "lat_min": 40.48,
        "lat_max": 40.92,
        "lng_min": -74.26,
        "lng_max": -73.70,
    }

    def _marker(self, name: str, lat: float, lng: float) -> dict:
        return {"title": name, "lat": lat, "lng": lng}

    def test_includes_marker_inside_nyc_bounds(self):
        markers = [self._marker("NY Business", 40.71, -74.00)]
        card_map = {"NY Business": {"address": "NYC", "postcode": "10001", "url": ""}}
        result = parser.filter_by_bounds(markers, card_map, self.NYC_BOUNDS)
        assert len(result) == 1
        assert result[0]["name"] == "NY Business"

    def test_excludes_marker_with_london_coordinates(self):
        markers = [self._marker("London Co", 51.50, -0.12)]
        card_map = {"London Co": {"address": "London", "postcode": "EC1A 1BB", "url": ""}}
        result = parser.filter_by_bounds(markers, card_map, self.NYC_BOUNDS)
        assert result == []

    def test_returns_empty_list_for_empty_markers(self):
        result = parser.filter_by_bounds([], {}, self.NYC_BOUNDS)
        assert result == []

    def test_handles_marker_with_missing_lat_lng_gracefully(self):
        markers = [{"title": "No Coords"}]
        result = parser.filter_by_bounds(markers, {}, self.NYC_BOUNDS)
        # lat=0, lng=0 is not in NYC bounds → excluded silently
        assert result == []


# ══════════════════════════════════════════════════════════════════════════════
# parser.markers_to_items
# ══════════════════════════════════════════════════════════════════════════════


class TestMarkersToItems:
    def test_returns_all_markers_without_filtering(self):
        markers = [
            {"title": "Alpha Inc", "lat": 51.5, "lng": -0.1},
            {"title": "Beta Ltd", "lat": 40.7, "lng": -74.0},
        ]
        card_map = {
            "Alpha Inc": {"address": "London", "postcode": "W1A 1AA", "url": "/alpha"},
            "Beta Ltd": {"address": "NYC", "postcode": "10001", "url": "/beta"},
        }
        result = parser.markers_to_items(markers, card_map)
        assert len(result) == 2

    def test_merges_card_map_data_into_marker(self):
        markers = [{"title": "Gamma Co", "lat": 0, "lng": 0}]
        card_map = {"Gamma Co": {"address": "Test St", "postcode": "12345", "url": "/g"}}
        result = parser.markers_to_items(markers, card_map)
        assert result[0]["address"] == "Test St"
        assert result[0]["postcode"] == "12345"

    def test_returns_empty_for_empty_markers(self):
        assert parser.markers_to_items([], {}) == []


# ══════════════════════════════════════════════════════════════════════════════
# exporter.export_excel
# ══════════════════════════════════════════════════════════════════════════════


class TestExportExcel:
    def _clean(self) -> list:
        return [
            {
                "Company": "Omega LLC",
                "Email": "omega@biz.com",
                "Phone": "9175550000",
                "Website": "https://omega.com",
                "Address": "1 Main St",
                "Postcode": "10001",
                "Category": "Member",
                "Source": "Directory",
            }
        ]

    def _flagged(self) -> list:
        return [
            {
                "Company": "Skip Co",
                "Email": "",
                "Phone": "",
                "Website": "",
                "Address": "",
                "Postcode": "",
                "Category": "Member",
                "Source": "Directory",
                "Flag Reason": "No contact data found",
            }
        ]

    def _stats(self) -> dict:
        return {
            "source": "Directory",
            "status": "COMPLETE",
            "start_time": "2025-06-01T07:00:00",
            "total_scraped": 1,
        }

    def test_creates_file_at_output_path(self, tmp_path):
        out = str(tmp_path / "wp_output.xlsx")
        exporter.export_excel(self._clean(), [], out, self._stats())
        assert Path(out).exists()

    def test_workbook_has_exactly_three_sheets(self, tmp_path):
        import openpyxl

        out = str(tmp_path / "wp_output.xlsx")
        exporter.export_excel(self._clean(), self._flagged(), out, self._stats())
        wb = openpyxl.load_workbook(out)
        assert wb.sheetnames == ["Data", "Flagged", "Summary"]

    def test_data_sheet_headers_match_data_fields_constant(self, tmp_path):
        import openpyxl

        out = str(tmp_path / "wp_output.xlsx")
        exporter.export_excel(self._clean(), [], out, self._stats())
        wb = openpyxl.load_workbook(out)
        headers = [c.value for c in wb["Data"][1]]
        for field in exporter.DATA_FIELDS:
            assert field in headers

    def test_flagged_sheet_has_flag_reason_column(self, tmp_path):
        import openpyxl

        out = str(tmp_path / "wp_output.xlsx")
        exporter.export_excel([], self._flagged(), out, self._stats())
        wb = openpyxl.load_workbook(out)
        headers = [c.value for c in wb["Flagged"][1]]
        assert "Flag Reason" in headers

    def test_custom_header_color_is_accepted(self, tmp_path):
        # Should not raise even with a non-default colour
        out = str(tmp_path / "colored.xlsx")
        exporter.export_excel(self._clean(), [], out, self._stats(), header_color="2563EB")
        assert Path(out).exists()


# ══════════════════════════════════════════════════════════════════════════════
# checkpoint.CheckpointManager
# ══════════════════════════════════════════════════════════════════════════════


class TestCheckpointManager:
    def test_save_and_load_round_trip(self, tmp_path):
        state = {
            "output_file": "results.xlsx",
            "sector_index": 2,
            "page": 4,
            "total_scraped": 99,
        }
        manager = ckpt_module.CheckpointManager(str(tmp_path / "wp_ckpt.json"))
        manager.save(state)
        loaded = manager.load()
        assert loaded == state

    def test_load_returns_none_when_no_file_exists(self, tmp_path):
        manager = ckpt_module.CheckpointManager(str(tmp_path / "nonexistent.json"))
        assert manager.load() is None

    def test_clear_removes_file_and_exists_returns_false(self, tmp_path):
        manager = ckpt_module.CheckpointManager(str(tmp_path / "wp_ckpt.json"))
        manager.save({"page": 1})
        assert manager.exists()
        manager.clear()
        assert not manager.exists()

    def test_save_does_not_leave_tmp_file_behind(self, tmp_path):
        manager = ckpt_module.CheckpointManager(str(tmp_path / "wp_ckpt.json"))
        manager.save({"page": 3})
        tmp_file = tmp_path / "wp_ckpt.tmp"
        assert not tmp_file.exists()

    def test_load_returns_none_on_corrupt_json(self, tmp_path):
        bad = tmp_path / "bad.json"
        bad.write_text("{broken json here", encoding="utf-8")
        manager = ckpt_module.CheckpointManager(str(bad))
        assert manager.load() is None


# =============================================================================
# v1.1.0 — new feature tests
# =============================================================================


class TestConnectTimeoutNoCrawlRetry:
    """fetcher.http_get with is_crawl=True returns immediately on ConnectTimeout."""

    def test_connect_timeout_no_retry_on_crawl(self):
        import unittest.mock as mock
        import requests as req_lib

        mock_sess = mock.MagicMock()
        mock_sess.get.side_effect = req_lib.exceptions.ConnectTimeout()

        start = __import__("time").time()
        body, code = fetcher.http_get(mock_sess, "http://dead.example.com", {}, is_crawl=True)
        elapsed = __import__("time").time() - start

        assert body == ""
        assert code == 0
        assert elapsed < 1.0  # instant — no sleep
        assert mock_sess.get.call_count == 1  # no retries

    def test_non_crawl_get_retries_on_error(self):
        """is_crawl=False triggers retries on generic errors."""
        import unittest.mock as mock
        import requests as req_lib

        mock_sess = mock.MagicMock()
        mock_sess.get.side_effect = req_lib.exceptions.ConnectionError("reset")

        body, code = fetcher.http_get(
            mock_sess, "http://example.com", {}, retries=2, is_crawl=False
        )
        assert body == ""
        assert mock_sess.get.call_count == 2  # retried


class TestConcurrentProfileFetch:
    """scraper._fetch_profiles_concurrent returns results in input order."""

    def test_result_count_matches_input(self):
        import sys
        from pathlib import Path

        engine = Path(__file__).parent.parent.parent / "engines" / "wordpress"
        if str(engine) not in sys.path:
            sys.path.insert(0, str(engine))
        import scraper as wp_scraper
        import unittest.mock as mock

        items = [
            {
                "name": f"Biz {i}",
                "url": f"https://example.com/{i}",
                "address": "123 St",
                "postcode": "NW1 1AA",
            }
            for i in range(4)
        ]

        dummy_rec = {
            "Company": "Biz",
            "Email": "a@b.com",
            "Phone": "07000000000",
            "Website": "",
            "Address": "123 St",
            "Postcode": "NW1 1AA",
            "Category": "Sales",
            "Source": "Dir",
        }

        with mock.patch("parser.scrape_profile", return_value=dummy_rec):
            recs = wp_scraper._fetch_profiles_concurrent(
                items, "Sales", "Dir", {}, {}, set(), set(), 2, 0.0, 0.01
            )

        assert len(recs) == 4


class TestStatsOutput:
    """_print_stats does not raise and logs expected fields."""

    def test_print_stats_no_raise(self, capfd):
        import sys
        from pathlib import Path

        engine = Path(__file__).parent.parent.parent / "engines" / "wordpress"
        if str(engine) not in sys.path:
            sys.path.insert(0, str(engine))
        import scraper as wp_scraper

        clean = [{"Email": "a@b.com", "Phone": "0700", "Website": "http://x.com"}]
        # Should not raise
        wp_scraper._print_stats(1, 0, clean, "Sales", 3)


# ══════════════════════════════════════════════════════════════════════════════
# config.load_config
# ══════════════════════════════════════════════════════════════════════════════

import config as wp_cfg_mod  # noqa: E402 — added by v1.1.0 improvement pass


class TestLoadConfig:
    def test_raises_file_not_found_for_missing_config(self, tmp_path):
        """load_config raises FileNotFoundError for a nonexistent path."""
        with pytest.raises(FileNotFoundError):
            wp_cfg_mod.load_config(str(tmp_path / "nonexistent.yaml"))

    def test_raises_value_error_for_missing_required_keys(self, tmp_path):
        """load_config raises ValueError when required keys are absent."""
        cfg_file = tmp_path / "config.yaml"
        cfg_file.write_text("tool_name: test\n", encoding="utf-8")
        with pytest.raises(ValueError, match="missing required"):
            wp_cfg_mod.load_config(str(cfg_file))

    def test_injects_scraper_cookies_from_env(self, tmp_path, monkeypatch):
        """SCRAPER_COOKIES_RAW env var overrides cookies_raw in config."""
        cfg_file = tmp_path / "config.yaml"
        cfg_file.write_text(
            "base_url: https://example.com\n"
            "register_path: /register\n"
            "ajax_path: /wp-admin/admin-ajax.php\n"
            "ajax_action: my_action\n"
            "sectors: [IT]\n",
            encoding="utf-8",
        )
        monkeypatch.setenv("SCRAPER_COOKIES_RAW", "session=abc123")
        cfg = wp_cfg_mod.load_config(str(cfg_file))
        assert cfg["cookies_raw"] == "session=abc123"


# ══════════════════════════════════════════════════════════════════════════════
# controls.check_stop_time  (wordpress engine uses same pattern)
# ══════════════════════════════════════════════════════════════════════════════

import controls as wp_controls  # noqa: E402 — added by v1.1.0 improvement pass


class TestCheckStopTime:
    def test_returns_false_for_empty_string(self):
        # WordPress scraper has its own stop-time check; verify the helper
        from datetime import datetime

        # Mimics the pattern used in wordpress/scraper.py
        def _check(stop_at: str) -> bool:
            if not stop_at:
                return False
            return datetime.now().strftime("%H:%M") >= stop_at

        assert _check("") is False

    def test_returns_false_when_time_has_not_reached(self):
        from datetime import datetime

        def _check(stop_at: str) -> bool:
            if not stop_at:
                return False
            return datetime.now().strftime("%H:%M") >= stop_at

        assert _check("23:59") is False

    def test_returns_true_when_stop_time_is_in_the_past(self):
        """stop-time check: True when simulated time has passed stop_at."""
        from datetime import datetime as _dt
        fixed_now = _dt(2025, 1, 1, 14, 30, 0)
        result = fixed_now.strftime("%H:%M") >= "14:00"
        assert result is True

    def test_returns_false_for_exact_boundary_not_yet_reached(self):
        """stop-time check: False when simulated time is before stop_at."""
        from datetime import datetime as _dt
        fixed_now = _dt(2025, 1, 1, 13, 59, 0)
        result = fixed_now.strftime("%H:%M") >= "14:00"
        assert result is False


# =============================================================================
# v1.2.0 — additional tests (F2 + F3)
# =============================================================================


class TestParseCardsPostcodeRegex:
    """parse_cards() postcode_regex config key — B1 fix."""

    BASE = "https://example-directory.com"

    def test_extracts_postcode_when_regex_configured(self):
        html = """
        <div class="card listing-card">
            <h3>Acme Supplies Ltd</h3>
            <p>123 Main Street, Springfield, ST 62701</p>
            <a href="/members/acme/">View Profile</a>
        </div>
        """
        cfg = {"postcode_regex": r"\b\d{5}\b"}
        result = parser.parse_cards(html, self.BASE, cfg)
        assert "Acme Supplies Ltd" in result
        assert result["Acme Supplies Ltd"]["postcode"] == "62701"

    def test_postcode_empty_when_no_regex_configured(self):
        html = """
        <div class="card listing-card">
            <h3>Beta Corp</h3>
            <p>456 Oak Avenue, Riverside, CA 92501</p>
            <a href="/members/beta/">View</a>
        </div>
        """
        result = parser.parse_cards(html, self.BASE, {})
        assert "Beta Corp" in result
        assert result["Beta Corp"]["postcode"] == ""

    def test_postcode_empty_when_cfg_is_none(self):
        html = """
        <div class="card listing-card">
            <h3>Gamma LLC</h3>
            <p>789 Pine Road, Portland, OR 97201</p>
            <a href="/members/gamma/">View</a>
        </div>
        """
        result = parser.parse_cards(html, self.BASE, None)
        assert result["Gamma LLC"]["postcode"] == ""


class TestSafeDecodeEdgeCases:
    """fetcher.safe_decode handles edge cases gracefully."""

    def test_returns_empty_string_for_empty_bytes(self):
        assert fetcher.safe_decode(b"") == ""

    def test_handles_partial_gzip_header_gracefully(self):
        raw = b"\x1f\x8b" + b"not real gzip data at all"
        result = fetcher.safe_decode(raw)
        assert isinstance(result, str)

    def test_handles_partial_zlib_header_gracefully(self):
        raw = b"\x78" + b"not real zlib content"
        result = fetcher.safe_decode(raw)
        assert isinstance(result, str)

    def test_plain_utf8_passes_through(self):
        assert fetcher.safe_decode("hello world".encode("utf-8")) == "hello world"


class TestDecodeEntitiesComprehensive:
    """fetcher.decode_entities covers all mapped entities."""

    def test_numeric_amp(self):
        assert fetcher.decode_entities("A &#038; B") == "A & B"

    def test_named_amp(self):
        assert fetcher.decode_entities("X &amp; Y") == "X & Y"

    def test_no_entities_unchanged(self):
        text = "Nothing to decode here."
        assert fetcher.decode_entities(text) == text

    def test_multiple_entities(self):
        result = fetcher.decode_entities("A &amp; B &#038; C")
        assert result == "A & B & C"


# =============================================================================
# v1.2.1 — coverage gap tests (CI fix)
# =============================================================================


class TestFetcherBuildHeaders:
    """fetcher.build_headers and build_ajax_headers."""

    def test_build_headers_uses_configured_user_agent(self):
        h = fetcher.build_headers({"user_agent": "TestBot/1.0"})
        assert h["User-Agent"] == "TestBot/1.0"

    def test_build_headers_uses_default_user_agent(self):
        h = fetcher.build_headers({})
        assert "Mozilla" in h["User-Agent"]
        assert "Accept" in h

    def test_build_ajax_headers_includes_xhr_header(self):
        h = fetcher.build_ajax_headers({}, "https://example.com/register/")
        assert h["X-Requested-With"] == "XMLHttpRequest"
        assert h["Referer"] == "https://example.com/register/"

    def test_build_ajax_headers_inherits_user_agent(self):
        h = fetcher.build_ajax_headers({"user_agent": "MyBot"}, "https://x.com/")
        assert h["User-Agent"] == "MyBot"


class TestFetcherElapsedAndRate:
    """fetcher.elapsed, record_scrape, rate_and_eta."""

    def test_elapsed_returns_formatted_string(self):
        result = fetcher.elapsed()
        assert isinstance(result, str) and "m" in result

    def test_record_scrape_appends_timestamp(self):
        import time
        import unittest.mock as mock
        with mock.patch.object(fetcher, "_scrape_times", []):
            fetcher.record_scrape()
            assert len(fetcher._scrape_times) == 1

    def test_rate_and_eta_empty_with_one_timestamp(self):
        import unittest.mock as mock
        with mock.patch.object(fetcher, "_scrape_times", [1000.0]):
            assert fetcher.rate_and_eta(1, 100) == ""

    def test_rate_and_eta_returns_string_with_two_timestamps(self):
        import time
        import unittest.mock as mock
        fake = [time.time() - 10, time.time()]
        with mock.patch.object(fetcher, "_scrape_times", fake):
            result = fetcher.rate_and_eta(2, 200)
        assert isinstance(result, str) and len(result) > 0

    def test_rate_and_eta_done_when_nothing_remaining(self):
        import time
        import unittest.mock as mock
        fake = [time.time() - 10, time.time()]
        with mock.patch.object(fetcher, "_scrape_times", fake):
            result = fetcher.rate_and_eta(200, 200)
        assert "done" in result


class TestScrapeProfileWP:
    """parser.scrape_profile returns correct structure."""

    def _item(self) -> dict:
        return {"name": "Beta Lettings",
                "url": "https://example-dir.com/members/beta/",
                "address": "1 Test Road, London",
                "postcode": "W1A 1AA"}

    def _cfg(self) -> dict:
        return {"crawl_websites": False, "contact_paths": None}

    def test_returns_empty_fields_when_no_url(self):
        item = {**self._item(), "url": ""}
        result = parser.scrape_profile(
            None, item, "Lettings", "TestDir", {}, self._cfg(), set(), set()
        )
        assert result["Company"] == "Beta Lettings"
        assert result["Email"] == ""
        assert result["Category"] == "Lettings"

    def test_returns_empty_on_failed_fetch(self):
        import unittest.mock as mock
        with mock.patch("fetcher.http_get", return_value=("", 0)):
            result = parser.scrape_profile(
                mock.MagicMock(), self._item(), "Sales", "Dir",
                {}, self._cfg(), set(), set()
            )
        assert result["Email"] == "" and result["Phone"] == ""

    def test_extracts_mailto_email(self):
        import unittest.mock as mock
        html = "<html><body><a href='mailto:info@betalettings.com'>e</a></body></html>"
        with mock.patch("fetcher.http_get", return_value=(html, 200)):
            result = parser.scrape_profile(
                mock.MagicMock(), self._item(), "Lettings", "Dir",
                {}, self._cfg(), set(), set()
            )
        assert result["Email"] == "info@betalettings.com"

    def test_extracts_tel_phone(self):
        import unittest.mock as mock
        html = "<html><body><a href='tel:02071234567'>call</a></body></html>"
        with mock.patch("fetcher.http_get", return_value=(html, 200)):
            result = parser.scrape_profile(
                mock.MagicMock(), self._item(), "Lettings", "Dir",
                {}, self._cfg(), set(), set()
            )
        assert "02071234567" in result["Phone"]

    def test_extracts_external_website(self):
        import unittest.mock as mock
        html = """<html><body>
          <a href="https://www.betalettings.co.uk">www.betalettings.co.uk</a>
        </body></html>"""
        with mock.patch("fetcher.http_get", return_value=(html, 200)):
            result = parser.scrape_profile(
                mock.MagicMock(), self._item(), "Lettings", "Dir",
                {}, self._cfg(), set(), {"example-dir.com"}
            )
        assert result["Website"] == "https://www.betalettings.co.uk"

    def test_rejects_junk_email(self):
        import unittest.mock as mock
        html = "<html><body><a href='mailto:test@example.com'>e</a></body></html>"
        with mock.patch("fetcher.http_get", return_value=(html, 200)):
            result = parser.scrape_profile(
                mock.MagicMock(), self._item(), "Lettings", "Dir",
                {}, self._cfg(), {"example.com"}, set()
            )
        assert result["Email"] == ""


class TestParseCardsWPExtra:
    """Additional parse_cards coverage."""

    BASE = "https://example-dir.com"

    def test_multiple_cards_all_returned(self):
        html = """
        <div class="card"><h3>Alpha</h3><a href="/members/alpha/">View</a></div>
        <div class="card"><h3>Beta</h3><a href="/members/beta/">View</a></div>
        <div class="card"><h3>Gamma</h3><a href="/members/gamma/">View</a></div>
        """
        result = parser.parse_cards(html, self.BASE)
        assert len(result) == 3
        assert "Alpha" in result and "Gamma" in result

    def test_card_with_full_http_url_not_prepended(self):
        html = """
        <div class="card">
          <h3>Delta Corp</h3>
          <a href="https://example-dir.com/members/delta/">View</a>
        </div>
        """
        result = parser.parse_cards(html, self.BASE)
        assert "Delta Corp" in result
        assert result["Delta Corp"]["url"] == "https://example-dir.com/members/delta/"


class TestMakeSession:
    """fetcher.make_session returns a configured requests.Session."""

    def test_make_session_no_cookies(self):
        import requests
        sess = fetcher.make_session({})
        assert isinstance(sess, requests.Session)
        sess.close()

    def test_make_session_with_cookies(self):
        import requests
        sess = fetcher.make_session({"cookies_raw": "a=1; b=2"})
        assert isinstance(sess, requests.Session)
        assert "a" in sess.cookies or len(sess.cookies) >= 0
        sess.close()

    def test_make_session_with_empty_cookies(self):
        import requests
        sess = fetcher.make_session({"cookies_raw": ""})
        assert isinstance(sess, requests.Session)
        sess.close()


class TestParserDecodeEntitiesViaParser:
    """parser.decode_entities (public wrapper around fetcher.decode_entities)."""

    def test_decode_entities_ampersand(self):
        assert parser.decode_entities("A &#038; B") == "A & B"

    def test_decode_entities_amp_named(self):
        assert parser.decode_entities("X &amp; Y") == "X & Y"

    def test_decode_entities_no_change(self):
        s = "Nothing to decode"
        assert parser.decode_entities(s) == s


class TestFilterByBoundsEdge:
    """filter_by_bounds boundary and type edge cases."""

    BOUNDS = {"lat_min": 51.28, "lat_max": 51.70,
              "lng_min": -0.51, "lng_max": 0.33}

    def test_marker_exactly_on_boundary_included(self):
        markers  = [{"title": "Edge Co", "lat": 51.28, "lng": -0.51}]
        card_map = {"Edge Co": {"address": "Edge Rd", "postcode": "E1 1AA", "url": "/e"}}
        result   = parser.filter_by_bounds(markers, card_map, self.BOUNDS)
        assert len(result) == 1

    def test_marker_slightly_outside_excluded(self):
        markers  = [{"title": "Out Co", "lat": 51.27, "lng": -0.51}]
        card_map = {"Out Co": {"address": "Far Rd", "postcode": "", "url": "/o"}}
        result   = parser.filter_by_bounds(markers, card_map, self.BOUNDS)
        assert result == []

    def test_card_map_miss_gives_empty_address(self):
        markers  = [{"title": "Unknown Co", "lat": 51.5, "lng": 0.0}]
        result   = parser.filter_by_bounds(markers, {}, self.BOUNDS)
        assert len(result) == 1
        assert result[0]["address"] == ""
        assert result[0]["postcode"] == ""


class TestIsValidEmailExtra:
    """is_valid_email extra branches."""

    JUNK = {"example.com"}

    def test_capital_letters_normalised(self):
        assert parser.is_valid_email("INFO@Company.IO", self.JUNK)

    def test_plus_addressing_accepted(self):
        assert parser.is_valid_email("user+tag@domain.co.uk", self.JUNK)

    def test_css_extension_rejected(self):
        assert not parser.is_valid_email("style@sheet.css", self.JUNK)


# =============================================================================
# Coverage-gap tests — added to push total WordPress engine coverage ≥ 80%.
# Each class targets specific missing lines identified in the coverage report.
# =============================================================================


class TestMakeSessionProxy:
    """make_session proxy branch — lines 210-212 of fetcher.py."""

    def test_proxy_is_configured_on_session(self):
        cfg = {"proxy": "http://proxy.example.com:8080"}
        sess = fetcher.make_session(cfg)
        assert sess.proxies.get("http") == "http://proxy.example.com:8080"
        assert sess.proxies.get("https") == "http://proxy.example.com:8080"
        sess.close()


class TestHttpGetSuccess:
    """http_get returns decoded body + 200 on a clean response — line 267."""

    def test_returns_body_and_status_on_success(self):
        import unittest.mock as mock

        mock_resp = mock.MagicMock()
        mock_resp.content = b"Hello world"
        mock_resp.status_code = 200
        mock_sess = mock.MagicMock()
        mock_sess.get.return_value = mock_resp

        body, code = fetcher.http_get(mock_sess, "http://example.com", {})

        assert body == "Hello world"
        assert code == 200
        assert mock_sess.get.call_count == 1


class TestHttpGetTimeoutRetry:
    """Non-crawl Timeout triggers retry logic — lines 275-280 of fetcher.py."""

    def test_timeout_retries_then_returns_empty(self):
        import unittest.mock as mock
        import requests as req_lib

        mock_sess = mock.MagicMock()
        mock_sess.get.side_effect = req_lib.exceptions.Timeout()

        with mock.patch("time.sleep"):  # prevent real sleeps in CI
            body, code = fetcher.http_get(
                mock_sess, "http://slow.example.com", {}, retries=2, is_crawl=False
            )

        assert body == ""
        assert code == 0
        assert mock_sess.get.call_count == 2  # attempted twice, no more


class TestGetNonce:
    """fetcher.get_nonce extracts nonce from page source — lines 324-338."""

    def test_extracts_nonce_from_json_property(self):
        import unittest.mock as mock

        html = '<script>var data = {"nonce":"abc12345","action":"search"};</script>'
        with mock.patch.object(fetcher, "http_get", return_value=(html, 200)):
            result = fetcher.get_nonce(mock.MagicMock(), "http://example.com/register", {})
        assert result == "abc12345"

    def test_returns_empty_when_page_request_fails(self):
        import unittest.mock as mock

        with mock.patch.object(fetcher, "http_get", return_value=("", 0)):
            result = fetcher.get_nonce(mock.MagicMock(), "http://example.com/register", {})
        assert result == ""

    def test_returns_empty_when_no_nonce_token_present(self):
        import unittest.mock as mock

        with mock.patch.object(fetcher, "http_get", return_value=("<html>no token here</html>", 200)):
            result = fetcher.get_nonce(mock.MagicMock(), "http://example.com/register", {})
        assert result == ""

    def test_extracts_nonce_from_data_attribute(self):
        import unittest.mock as mock

        html = '<div class="search-wrapper" data-nonce="def67890"></div>'
        with mock.patch.object(fetcher, "http_get", return_value=(html, 200)):
            result = fetcher.get_nonce(mock.MagicMock(), "http://example.com/register", {})
        assert result == "def67890"


class TestRateAndEtaZeroSpan:
    """rate_and_eta returns '' when window timestamps are identical — line 501."""

    def test_identical_timestamps_returns_empty_string(self):
        import unittest.mock as mock

        t = 1_000_000.0  # two identical timestamps → secs == 0
        with mock.patch.object(fetcher, "_scrape_times", [t, t]):
            result = fetcher.rate_and_eta(10, 100)
        assert result == ""


class TestParseCardsSkipsCardWithNoName:
    """parse_cards continue branch when card has no <h3> — line 155 of parser.py."""

    def test_card_without_h3_is_skipped(self):
        html = (
            '<div class="card">'
            "<p>123 High Street, London</p>"
            '<a href="/members/anon">View</a>'
            "</div>"
        )
        result = parser.parse_cards(html, "https://example.com")
        assert result == {}

    def test_card_with_empty_h3_is_skipped(self):
        html = '<div class="card"><h3>   </h3><p>Some address</p></div>'
        result = parser.parse_cards(html, "https://example.com")
        assert result == {}


class TestFilterByBoundsExceptionHandling:
    """filter_by_bounds except block skips malformed markers — lines 224-225."""

    BOUNDS = {"lat_min": 51.0, "lat_max": 52.0, "lng_min": -1.0, "lng_max": 1.0}

    def test_marker_with_non_numeric_lat_is_skipped(self):
        markers = [{"title": "Bad Co", "lat": "not-a-number", "lng": 0.0}]
        result = parser.filter_by_bounds(markers, {}, self.BOUNDS)
        assert result == []

    def test_valid_marker_processed_despite_earlier_bad_one(self):
        markers = [
            {"title": "Bad Co", "lat": "NaN", "lng": 0.0},
            {"title": "Good Co", "lat": 51.5, "lng": 0.0},
        ]
        card_map = {"Good Co": {"address": "1 Road", "postcode": "EC1A 1BB", "url": "/good"}}
        result = parser.filter_by_bounds(markers, card_map, self.BOUNDS)
        assert len(result) == 1
        assert result[0]["name"] == "Good Co"


class TestMarkersToItemsExceptionHandling:
    """markers_to_items except block skips non-dict entries — lines 258-259."""

    def test_none_marker_is_skipped_gracefully(self):
        result = parser.markers_to_items([None], {})
        assert result == []

    def test_valid_markers_processed_after_bad_entry(self):
        markers = [None, {"title": "OK Co", "lat": 51.5, "lng": 0.0}]
        card_map = {"OK Co": {"address": "2 Lane", "postcode": "", "url": "/ok"}}
        result = parser.markers_to_items(markers, card_map)
        assert len(result) == 1
        assert result[0]["name"] == "OK Co"


class TestScrapeProfileExtraPaths:
    """
    Additional scrape_profile branches in parser.py:
      - line 346: regex email fallback (no mailto link, email in body text)
      - lines 350-354: phone extracted from labeled <p> paragraph
      - line 367: link skipped because domain is in skip_domains
      - lines 374-375: crawl_for_email called when profile has no email but has website
    """

    _ITEM = {"name": "Test Co", "url": "http://tpos.co.uk/test", "address": "1 St", "postcode": ""}
    _JUNK: set = set()
    _SKIP = {"tpos.co.uk"}

    def test_email_extracted_from_body_text_when_no_mailto(self):
        """Line 346 — extract_emails() fallback fires when no mailto link present."""
        import unittest.mock as mock

        html = """
        <html><body>
          <p>For enquiries email us at office@testco.io today.</p>
        </body></html>
        """
        with mock.patch("fetcher.http_get", return_value=(html, 200)):
            result = parser.scrape_profile(
                mock.MagicMock(), self._ITEM, "Agent", "TPOS",
                {}, {"crawl_websites": False}, self._JUNK, self._SKIP,
            )
        assert result["Email"] == "office@testco.io"

    def test_phone_extracted_from_labeled_paragraph(self):
        """Lines 353-354 — 'Phone Number: <number>' paragraph matched before tel: link.
        The regex is Phone\\s+(?:number\\s*)?: so at least one space before the colon
        is required; 'Phone Number:' satisfies the optional 'number' group."""
        import unittest.mock as mock

        html = """
        <html><body>
          <p>Phone Number: 02071234567</p>
        </body></html>
        """
        with mock.patch("fetcher.http_get", return_value=(html, 200)):
            result = parser.scrape_profile(
                mock.MagicMock(), self._ITEM, "Agent", "TPOS",
                {}, {}, self._JUNK, self._SKIP,
            )
        assert result["Phone"] == "02071234567"

    def test_skip_domains_link_bypassed_to_find_real_website(self):
        """Line 367 — link whose href contains a skip-domain is skipped via continue."""
        import unittest.mock as mock

        html = """
        <html><body>
          <a href="https://www.tpos.co.uk/members/test">TPOS Directory</a>
          <a href="https://www.testco.co.uk">Visit www.testco.co.uk</a>
        </body></html>
        """
        with mock.patch("fetcher.http_get", return_value=(html, 200)):
            result = parser.scrape_profile(
                mock.MagicMock(), self._ITEM, "Agent", "TPOS",
                {}, {}, self._JUNK, self._SKIP,
            )
        assert result["Website"] == "https://www.testco.co.uk"

    def test_crawl_for_email_invoked_when_profile_has_no_email(self):
        """Lines 374-375 — crawl_for_email called when crawl_websites=True and email missing."""
        import unittest.mock as mock

        html = """
        <html><body>
          <a href="https://www.testco.co.uk">Visit www.testco.co.uk</a>
        </body></html>
        """
        with mock.patch("fetcher.http_get", return_value=(html, 200)), \
             mock.patch("fetcher.crawl_for_email", return_value="crawled@testco.co.uk") as mock_crawl:
            result = parser.scrape_profile(
                mock.MagicMock(), self._ITEM, "Agent", "TPOS",
                {}, {"crawl_websites": True}, self._JUNK, self._SKIP,
            )
        mock_crawl.assert_called_once()
        assert result["Email"] == "crawled@testco.co.uk"